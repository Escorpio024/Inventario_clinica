from fastapi import FastAPI, Depends, HTTPException, status, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy import create_engine, text, func
from sqlalchemy.orm import sessionmaker, Session, joinedload
from contextlib import asynccontextmanager
from datetime import datetime, timedelta, date
from passlib.context import CryptContext
from jose import JWTError, jwt
from fastapi.security import OAuth2PasswordBearer
from openpyxl import Workbook
from io import BytesIO
import os
import logging

from .models import Base, User, Product, Lot, Movement, MovementType, UserRole, Empresa
from .schemas import (
    UserLogin, Token, UserCreate, UserUpdate, UserOut,
    ProductCreate, ProductOut, LotCreate, LotOut, MovementCreate, MovementOut,
    EmpresaCreate, EmpresaOut, SaaSRegister
)

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ── Database ──────────────────────────────────────────────────────────────────
DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://postgres:postgres@localhost/inventario")
engine = create_engine(
    DATABASE_URL,
    pool_size=10,
    max_overflow=20,
    pool_pre_ping=True,
    pool_recycle=300,
)
SessionLocal = sessionmaker(bind=engine, autocommit=False, autoflush=False)

# ── Security ──────────────────────────────────────────────────────────────────
SECRET_KEY = os.getenv("SECRET_KEY")
if not SECRET_KEY:
    raise RuntimeError("SECRET_KEY environment variable is not set!")

ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 480  # 8 horas

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/auth/login")

# ── CORS ──────────────────────────────────────────────────────────────────────
ALLOWED_ORIGINS = os.getenv("ALLOWED_ORIGINS", "http://localhost:3000").split(",")

# ── Lifespan ──────────────────────────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    logger.info("🚀 Iniciando servidor...")
    Base.metadata.create_all(bind=engine)
    _migrate_users_table()
    _migrate_multitenant()       # <-- nueva migración multi-tenant
    _migrate_new_fields()        # <-- migración para refactor Lote/Producto
    _create_default_empresa()    # <-- crea Empresa 1 si no existe
    _create_default_admin()
    logger.info("✅ Servidor listo.")
    yield
    logger.info("🛑 Cerrando servidor.")

def _migrate_users_table():
    """Agrega columnas nuevas a la tabla users si no existen (safe migration)."""
    with engine.connect() as conn:
        for col_sql in [
            "ALTER TABLE users ADD COLUMN full_name VARCHAR",
            "ALTER TABLE users ADD COLUMN role VARCHAR DEFAULT 'OPERADOR'",
            "ALTER TABLE users ADD COLUMN created_at TIMESTAMP DEFAULT NOW()",
        ]:
            try:
                conn.execute(text(col_sql))
                conn.commit()
            except Exception:
                conn.rollback()
        # Migración segura del enum en PostgreSQL: agrega 'SUPERADMIN' si no existe
        try:
            conn.execute(text("ALTER TYPE userrole ADD VALUE IF NOT EXISTS 'SUPERADMIN'"))
            conn.commit()
        except Exception:
            conn.rollback()  # SQLite o ya existe, ignorar
        try:
            conn.execute(text("UPDATE users SET role = 'OPERADOR' WHERE role IS NULL"))
            conn.execute(text("UPDATE users SET created_at = NOW() WHERE created_at IS NULL"))
            # El admin principal es ahora SUPERADMIN
            conn.execute(text("UPDATE users SET role = 'SUPERADMIN', full_name = 'Super Administrador' WHERE email = 'admin@clinica.com'"))
            conn.commit()
        except Exception:
            conn.rollback()

def _migrate_multitenant():
    """Agrega columna empresa_id a todas las tablas de negocio si no existe."""
    tables = ["users", "products", "lots", "movements"]
    with engine.connect() as conn:
        for table in tables:
            try:
                conn.execute(text(f"ALTER TABLE {table} ADD COLUMN empresa_id INTEGER REFERENCES empresas(id)"))
                conn.commit()
                logger.info(f"✅ empresa_id agregado a {table}")
            except Exception:
                conn.rollback()  # Ya existe, ignorar
        # Asignar empresa_id=1 a todos los registros existentes sin empresa
        for table in tables:
            try:
                conn.execute(text(f"UPDATE {table} SET empresa_id = 1 WHERE empresa_id IS NULL"))
                conn.commit()
            except Exception:
                conn.rollback()

def _migrate_new_fields():
    """Agrega columnas nuevas a products y lots si no existen."""
    product_cols = [
        "presentacion VARCHAR", "registro_sanitario VARCHAR",
        "principio_activo VARCHAR", "forma_farmaceutica VARCHAR",
        "concentracion VARCHAR", "marca VARCHAR",
        "vida_util VARCHAR", "clasificacion_riesgo VARCHAR"
    ]
    lot_cols = [
        "factura VARCHAR", "fecha_recepcion DATE",
        "estado_recepcion VARCHAR", "causas_rechazo VARCHAR"
    ]
    mov_cols = [
        "patient VARCHAR", "doctor VARCHAR", "destination VARCHAR"
    ]
    with engine.connect() as conn:
        for col in product_cols:
            try:
                conn.execute(text(f"ALTER TABLE products ADD COLUMN {col}"))
                conn.commit()
            except Exception:
                conn.rollback()
        for col in lot_cols:
            try:
                conn.execute(text(f"ALTER TABLE lots ADD COLUMN {col}"))
                conn.commit()
            except Exception:
                conn.rollback()
        for col in mov_cols:
            try:
                conn.execute(text(f"ALTER TABLE movements ADD COLUMN {col}"))
                conn.commit()
            except Exception:
                conn.rollback()

def _create_default_empresa():
    """Crea la empresa 1 (clínica demo) si no existe."""
    db = SessionLocal()
    try:
        empresa = db.query(Empresa).filter(Empresa.id == 1).first()
        if not empresa:
            empresa = Empresa(
                nombre="Clínica Demo",
                slug="clinica-demo",
                plan="PRO",
            )
            db.add(empresa)
            db.commit()
            logger.info("✅ Empresa demo creada (ID=1).")
    finally:
        db.close()


def _create_default_admin():
    db = SessionLocal()
    try:
        admin = db.query(User).filter(User.email == "admin@clinica.com").first()
        if not admin:
            hashed = pwd_context.hash("admin123")
            admin = User(
                email="admin@clinica.com",
                hashed_password=hashed,
                full_name="Super Administrador",
                role=UserRole.SUPERADMIN,
                empresa_id=1,
            )
            db.add(admin)
            db.commit()
            logger.info("✅ Super Admin creado (empresa_id=1).")
        else:
            # Promover a SUPERADMIN si era ADMIN
            if admin.role != UserRole.SUPERADMIN:
                admin.role = UserRole.SUPERADMIN
                admin.full_name = "Super Administrador"
            if not admin.empresa_id:
                admin.empresa_id = 1
            db.commit()
    finally:
        db.close()

# ── App ───────────────────────────────────────────────────────────────────────
app = FastAPI(
    title="Inventario Clínica API",
    version="1.2.0",
    description="API REST para gestión de inventario médico con trazabilidad FEFO y roles de usuario",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Health Check (para UptimeRobot — evita Cold Start en Render) ─────────────
@app.get("/health", tags=["system"], include_in_schema=False)
def health_check():
    """Endpoint ultraligero para mantener el servidor de Render activo.
    Configura UptimeRobot para hacer ping aquí cada 5 minutos."""
    return {"status": "ok"}

# ── Dependencies ──────────────────────────────────────────────────────────────
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def get_current_user(
    token: str = Depends(oauth2_scheme),
    db: Session = Depends(get_db)
) -> User:
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="No se pudo validar las credenciales",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        if email is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception

    user = db.query(User).filter(User.email == email, User.is_active == 1).first()
    if user is None:
        raise credentials_exception
    return user

def get_empresa_id(user: User = Depends(get_current_user)) -> int:
    """Retorna el empresa_id del usuario autenticado (con fallback a 1)."""
    return user.empresa_id or 1

def require_admin(user: User = Depends(get_current_user)) -> User:
    """Requiere ADMIN o SUPERADMIN."""
    if user.role not in (UserRole.ADMIN, UserRole.SUPERADMIN):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="❌ Acceso denegado. Se requiere rol de Administrador."
        )
    return user

def require_superadmin(user: User = Depends(get_current_user)) -> User:
    """Requiere SUPERADMIN exclusivamente."""
    if user.role != UserRole.SUPERADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="❌ Acceso denegado. Solo el Super Administrador puede hacer esto."
        )
    return user

# ── Auth ──────────────────────────────────────────────────────────────────────
@app.post("/auth/login", response_model=Token, tags=["auth"])
def login(credentials: UserLogin, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == credentials.email, User.is_active == 1).first()
    if not user or not pwd_context.verify(credentials.password, user.hashed_password):
        raise HTTPException(status_code=400, detail="Credenciales incorrectas o usuario inactivo")

    access_token = jwt.encode(
        {
            "sub": user.email,
            "empresa_id": user.empresa_id or 1,
            "exp": datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
        },
        SECRET_KEY,
        algorithm=ALGORITHM
    )
    empresa_nombre = None
    if user.empresa_id:
        empresa = db.query(Empresa).filter(Empresa.id == user.empresa_id).first()
        empresa_nombre = empresa.nombre if empresa else None
    logger.info(f"🔐 Login exitoso: {user.email} (empresa_id: {user.empresa_id}, rol: {user.role})")
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "role": user.role.value,
        "full_name": user.full_name or user.email,
        "empresa_id": user.empresa_id or 1,
        "empresa_nombre": empresa_nombre,
    }

@app.post("/auth/register", response_model=Token, tags=["auth"])
def register_saas(payload: SaaSRegister, db: Session = Depends(get_db)):
    import re
    # Check if email exists in any company
    if db.query(User).filter(User.email == payload.admin_email).first():
        raise HTTPException(status_code=400, detail="El correo ya está registrado en el sistema")
    
    # Create Empresa
    slug = re.sub(r'[^a-z0-9]+', '-', payload.empresa_nombre.lower()).strip('-')
    if db.query(Empresa).filter(Empresa.slug == slug).first():
        slug = f"{slug}-{int(datetime.utcnow().timestamp())}"
        
    empresa = Empresa(nombre=payload.empresa_nombre, slug=slug, plan="FREE")
    db.add(empresa)
    db.commit()
    db.refresh(empresa)
    
    # Create Admin User
    hashed = pwd_context.hash(payload.admin_password)
    user = User(
        email=payload.admin_email,
        hashed_password=hashed,
        full_name=payload.admin_name,
        role=UserRole.ADMIN,
        empresa_id=empresa.id
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    
    logger.info(f"🚀 NUEVO CLIENTE SAAS: {empresa.nombre} por {user.email}")
    
    # Auto-login to return token immediately
    access_token = jwt.encode(
        {
            "sub": user.email,
            "empresa_id": empresa.id,
            "exp": datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
        },
        SECRET_KEY,
        algorithm=ALGORITHM
    )
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "role": user.role.value,
        "full_name": user.full_name,
        "empresa_id": empresa.id,
        "empresa_nombre": empresa.nombre,
    }

@app.get("/auth/me", response_model=UserOut, tags=["auth"])
def get_me(user: User = Depends(get_current_user)):
    return user


# ── Empresas (solo SUPERADMIN) ───────────────────────────────────────────────────
@app.get("/empresas", response_model=list[EmpresaOut], tags=["empresas"])
def list_empresas(db: Session = Depends(get_db), sa: User = Depends(require_superadmin)):
    """Lista todas las empresas. Solo SUPERADMIN."""
    return db.query(Empresa).order_by(Empresa.id).all()

@app.post("/empresas", response_model=EmpresaOut, status_code=status.HTTP_201_CREATED, tags=["empresas"])
def create_empresa(payload: EmpresaCreate, db: Session = Depends(get_db), sa: User = Depends(require_superadmin)):
    import re
    # Verificar si ya existe una empresa con ese nombre
    if db.query(Empresa).filter(Empresa.nombre == payload.nombre).first():
        raise HTTPException(400, f"Ya existe una empresa llamada '{payload.nombre}'")
    
    # Usar el slug del payload o auto-generarlo desde el nombre
    slug = payload.slug if payload.slug else re.sub(r'[^a-z0-9]+', '-', payload.nombre.lower()).strip('-')
    # Si el slug ya existe, añadir timestamp para hacerlo único
    if db.query(Empresa).filter(Empresa.slug == slug).first():
        slug = f"{slug}-{int(datetime.utcnow().timestamp())}"
        
    empresa = Empresa(nombre=payload.nombre, slug=slug, plan=payload.plan)
    db.add(empresa)
    db.commit()
    db.refresh(empresa)
    logger.info(f"🏢 Empresa '{empresa.nombre}' (slug: {slug}) creada por {sa.email}")
    return empresa


# ── Dashboard Stats ───────────────────────────────────────────────────────────
@app.get("/stats", tags=["dashboard"])
def get_dashboard_stats(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    eid: int = Depends(get_empresa_id)
):    
    total_products = db.query(Product).filter(Product.empresa_id == eid).count()
    total_lots = db.query(Lot).filter(Lot.empresa_id == eid).count()
    vencidos_count = db.query(Lot).filter(Lot.empresa_id == eid, Lot.expiry_date < date.today()).count()
    
    # Pre-calcular también por redimiento
    total_valor = db.query(func.sum(Lot.qty_current * Lot.unit_cost)).filter(Lot.empresa_id == eid).scalar() or 0
    total_movements = db.query(Movement).filter(Movement.empresa_id == eid).count()

    return {
        "total_products": total_products,
        "total_lots": total_lots,
        "vencidos_count": vencidos_count,
        "valor_inventario": total_valor,
        "total_movements": total_movements
    }

# ── User Management ─────────────────────────────────────────────────────────────────
@app.get("/users", response_model=list[UserOut], tags=["users"])
def list_users(
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
    eid: int = Depends(get_empresa_id)
):
    # SUPERADMIN ve todos; ADMIN solo su empresa
    if admin.role == UserRole.SUPERADMIN:
        return db.query(User).order_by(User.empresa_id, User.created_at.desc()).all()
    return db.query(User).filter(User.empresa_id == eid).order_by(User.created_at.desc()).all()

@app.post("/users", response_model=UserOut, status_code=status.HTTP_201_CREATED, tags=["users"])
def create_user(
    payload: UserCreate,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
    eid: int = Depends(get_empresa_id)
):
    existing = db.query(User).filter(User.email == payload.email).first()
    if existing:
        raise HTTPException(400, f"Ya existe un usuario con el email '{payload.email}'")
    
    # Solo SUPERADMIN puede asignar el usuario a otra empresa
    if admin.role == UserRole.SUPERADMIN and payload.empresa_id:
        target_eid = payload.empresa_id
    else:
        target_eid = eid  # ADMIN hereda su propia empresa
    
    hashed = pwd_context.hash(payload.password)
    user = User(
        email=payload.email,
        hashed_password=hashed,
        full_name=payload.full_name,
        role=payload.role,
        empresa_id=target_eid,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    logger.info(f"👤 Usuario creado por {admin.email} (empresa {target_eid}): {user.email}")
    return user

@app.put("/users/{id}", response_model=UserOut, tags=["users"])
def update_user(
    id: int,
    payload: UserUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
    eid: int = Depends(get_empresa_id)
):
    # SUPERADMIN puede editar cualquier usuario; ADMIN solo los de su empresa
    if admin.role == UserRole.SUPERADMIN:
        user = db.query(User).filter(User.id == id).first()
    else:
        user = db.query(User).filter(User.id == id, User.empresa_id == eid).first()
    if not user:
        raise HTTPException(404, "Usuario no encontrado")
    # Evitar que se degrade el propio SUPERADMIN
    if user.email == admin.email and payload.role and payload.role not in (UserRole.ADMIN, UserRole.SUPERADMIN):
        raise HTTPException(400, "No puedes cambiar tu propio rol de Administrador")
    if payload.full_name is not None:
        user.full_name = payload.full_name
    if payload.role is not None:
        user.role = payload.role
    if payload.is_active is not None:
        if user.email == admin.email and payload.is_active == 0:
            raise HTTPException(400, "No puedes desactivar tu propia cuenta")
        user.is_active = payload.is_active
    if payload.password:
        user.hashed_password = pwd_context.hash(payload.password)
    db.commit()
    db.refresh(user)
    logger.info(f"✏️ Usuario {user.email} actualizado por {admin.email}")
    return user

@app.delete("/users/{id}", tags=["users"])
def delete_user(
    id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
    eid: int = Depends(get_empresa_id)
):
    # SUPERADMIN puede eliminar cualquier usuario; ADMIN solo los de su empresa
    if admin.role == UserRole.SUPERADMIN:
        user = db.query(User).filter(User.id == id).first()
    else:
        user = db.query(User).filter(User.id == id, User.empresa_id == eid).first()
    if not user:
        raise HTTPException(404, "Usuario no encontrado")
    if user.email == admin.email:
        raise HTTPException(400, "No puedes eliminar tu propia cuenta")
    db.delete(user)
    db.commit()
    logger.info(f"🗑️ Usuario {user.email} eliminado por {admin.email}")
    return {"message": f"Usuario '{user.email}' eliminado correctamente"}

# ── Products ──────────────────────────────────────────────────────────────────
@app.get("/products", response_model=list[ProductOut], tags=["products"])
def get_products(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    search: str = Query(None, max_length=100),
    category: str = Query(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    eid: int = Depends(get_empresa_id)
):
    query = db.query(Product).filter(Product.empresa_id == eid)
    if search:
        query = query.filter(
            Product.name.ilike(f"%{search}%") |
            Product.barcode.ilike(f"%{search}%")
        )
    if category:
        query = query.filter(Product.category == category)
    return query.order_by(Product.name).offset(skip).limit(limit).all()

@app.post("/products", response_model=ProductOut, status_code=status.HTTP_201_CREATED, tags=["products"])
def create_product(
    product: ProductCreate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    eid: int = Depends(get_empresa_id)
):
    existing = db.query(Product).filter(Product.empresa_id == eid, Product.name == product.name).first()
    if existing:
        raise HTTPException(400, f"Ya existe un producto con el nombre '{product.name}'")
    db_product = Product(**product.dict(), empresa_id=eid)
    db.add(db_product)
    db.commit()
    db.refresh(db_product)
    logger.info(f"📦 Producto '{product.name}' creado por {user.email} (empresa {eid})")
    return db_product

@app.delete("/products/{id}", tags=["products"])
def delete_product(
    id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    eid: int = Depends(get_empresa_id)
):
    if user.role != UserRole.ADMIN:
        raise HTTPException(403, "Solo los administradores pueden eliminar productos")
    product = db.query(Product).filter(Product.id == id, Product.empresa_id == eid).first()
    if not product:
        raise HTTPException(404, "Producto no encontrado")
    if db.query(Lot).filter(Lot.product_id == id).count() > 0:
        raise HTTPException(400, "No se puede eliminar: El producto tiene lotes asociados.")
    if db.query(Movement).filter(Movement.product_id == id).count() > 0:
        raise HTTPException(400, "No se puede eliminar: El producto tiene histórico de movimientos.")
    db.delete(product)
    db.commit()
    logger.info(f"🗑️ Producto '{product.name}' eliminado por {user.email}")
    return {"message": "Producto eliminado correctamente"}

# ── Lots ──────────────────────────────────────────────────────────────────────
@app.get("/lots", response_model=list[LotOut], tags=["lots"])
def get_lots(
    skip: int = Query(0, ge=0),
    limit: int = Query(200, ge=1, le=1000),
    search: str = Query(None, max_length=100),
    status_filter: str = Query(None, alias="status"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    eid: int = Depends(get_empresa_id)
):
    query = db.query(Lot).join(Product, Lot.product_id == Product.id, isouter=True) \
                         .filter(Lot.empresa_id == eid)
    if search:
        query = query.filter(
            Product.name.ilike(f"%{search}%") |
            Lot.lot_number.ilike(f"%{search}%") |
            Lot.barcode.ilike(f"%{search}%") |
            Lot.factura.ilike(f"%{search}%") |
            Product.category.ilike(f"%{search}%") |
            Product.principio_activo.ilike(f"%{search}%") |
            Product.marca.ilike(f"%{search}%")
        )
    if status_filter == "DANGER":
        query = query.filter(Lot.expiry_date <= date.today() + timedelta(days=90))
    elif status_filter == "WARNING":
        query = query.filter(
            Lot.expiry_date > date.today() + timedelta(days=90),
            Lot.expiry_date <= date.today() + timedelta(days=365)
        )
    elif status_filter == "OK":
        query = query.filter(Lot.expiry_date > date.today() + timedelta(days=365))
    return query.order_by(Lot.expiry_date.asc()).offset(skip).limit(limit).all()

@app.post("/lots", response_model=LotOut, status_code=status.HTTP_201_CREATED, tags=["lots"])
def create_lot(
    lot: LotCreate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    eid: int = Depends(get_empresa_id)
):
    product = db.query(Product).filter(Product.id == lot.product_id, Product.empresa_id == eid).first()
    if not product:
        raise HTTPException(400, "Producto no encontrado")

    existing = db.query(Lot).filter(
        Lot.product_id == lot.product_id,
        Lot.lot_number == lot.lot_number,
        Lot.empresa_id == eid
    ).first()
    if existing:
        raise HTTPException(400, f"Ya existe el lote '{lot.lot_number}' para este producto")

    db_lot = Lot(**lot.dict(), qty_current=lot.qty_initial, empresa_id=eid)
    db.add(db_lot)
    db.commit()
    db.refresh(db_lot)
    logger.info(f"🏷️ Lote '{lot.lot_number}' creado por {user.email} (empresa {eid})")
    return db_lot

@app.delete("/lots/{id}", tags=["lots"])
def delete_lot(
    id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    eid: int = Depends(get_empresa_id)
):
    if user.role != UserRole.ADMIN:
        raise HTTPException(403, "Solo los administradores pueden eliminar lotes")
    lot = db.query(Lot).filter(Lot.id == id, Lot.empresa_id == eid).first()
    if not lot:
        raise HTTPException(404, "Lote no encontrado")
    if db.query(Movement).filter(Movement.lot_id == id).count() > 0:
        raise HTTPException(400, "No se puede eliminar: El lote tiene movimientos asociados.")
    db.delete(lot)
    db.commit()
    return {"message": "Lote eliminado correctamente"}

@app.put("/lots/{id}", response_model=LotOut, tags=["lots"])
def update_lot(
    id: int,
    payload: LotCreate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    eid: int = Depends(get_empresa_id)
):
    if user.role not in (UserRole.ADMIN, UserRole.SUPERADMIN):
        raise HTTPException(403, "Solo los administradores pueden editar lotes")
    lot = db.query(Lot).filter(Lot.id == id, Lot.empresa_id == eid).first()
    if not lot:
        raise HTTPException(404, "Lote no encontrado")
    lot.lot_number = payload.lot_number
    lot.expiry_date = payload.expiry_date
    lot.unit_cost = payload.unit_cost
    lot.factura = payload.factura
    lot.fecha_recepcion = payload.fecha_recepcion
    lot.estado_recepcion = payload.estado_recepcion
    lot.causas_rechazo = payload.causas_rechazo
    lot.barcode = payload.barcode
    db.commit()
    db.refresh(lot)
    logger.info(f"✏️ Lote '{lot.lot_number}' actualizado por {user.email}")
    return lot

# ── Movements ─────────────────────────────────────────────────────────────────
@app.get("/movements", response_model=list[MovementOut], tags=["movements"])
def get_movements(
    skip: int = Query(0, ge=0),
    limit: int = Query(200, ge=1, le=1000),
    search: str = Query(None, max_length=100),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    eid: int = Depends(get_empresa_id)
):
    query = db.query(Movement).options(joinedload(Movement.product), joinedload(Movement.lot)) \
                              .filter(Movement.empresa_id == eid)
    
    if search:
        query = query.join(Product, Movement.product_id == Product.id, isouter=True) \
                     .join(Lot, Movement.lot_id == Lot.id, isouter=True) \
                     .filter(
                         Product.name.ilike(f"%{search}%") |
                         Lot.lot_number.ilike(f"%{search}%") |
                         Movement.reason.ilike(f"%{search}%") |
                         Movement.user_email.ilike(f"%{search}%")
                     )

    movements = query.order_by(Movement.created_at.desc()).offset(skip).limit(limit).all()
    return [
        {
            "id": m.id,
            "type": m.type.value,
            "product_id": m.product_id,
            "lot_id": m.lot_id,
            "qty": m.qty,
            "reason": m.reason,
            "patient": m.patient,
            "doctor": m.doctor,
            "destination": m.destination,
            "user_email": m.user_email,
            "created_at": m.created_at,
            "product_name": m.product.name if m.product else None,
            "lot_number": m.lot.lot_number if m.lot else None,
        }
        for m in movements
    ]

@app.post("/movements", status_code=status.HTTP_201_CREATED, tags=["movements"])
def create_movement(
    payload: MovementCreate,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    eid: int = Depends(get_empresa_id)
):
    if payload.qty <= 0:
        raise HTTPException(400, "La cantidad debe ser mayor a 0")

    product = db.query(Product).filter_by(id=payload.product_id, empresa_id=eid).first()
    if not product:
        raise HTTPException(400, "Producto inválido")

    lot = db.query(Lot).filter_by(id=payload.lot_id, empresa_id=eid).first()
    if not lot or lot.product_id != payload.product_id:
        raise HTTPException(400, "Lote inválido para el producto")

    if payload.type == MovementType.SALIDA and lot.expiry_date <= date.today():
        raise HTTPException(403, "❌ Lote vencido. No se puede usar para SALIDA.")

    if payload.type in (MovementType.SALIDA, MovementType.MERMA):
        if lot.qty_current < payload.qty:
            raise HTTPException(400, f"Stock insuficiente. Disponible: {lot.qty_current}")
        lot.qty_current -= payload.qty
    else:
        lot.qty_current += payload.qty

    movement = Movement(
        type=payload.type,
        product_id=payload.product_id,
        lot_id=payload.lot_id,
        qty=abs(payload.qty),
        reason=payload.reason,
        patient=payload.patient,
        doctor=payload.doctor,
        destination=payload.destination,
        user_email=user.email,
        empresa_id=eid,
    )
    db.add(movement)
    db.commit()
    logger.info(f"📋 Movimiento {payload.type.value} por {user.email}: {payload.qty} u. lote {lot.lot_number}")
    return {"status": "success", "message": "Movimiento registrado", "qty_current": lot.qty_current}

# ── Barcode search ────────────────────────────────────────────────────────────
@app.get("/search/barcode/{code}", tags=["search"])
def search_barcode(
    code: str,
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
    eid: int = Depends(get_empresa_id)
):
    lot = db.query(Lot).filter(Lot.barcode == code, Lot.empresa_id == eid).first()
    if lot:
        product = db.query(Product).filter(Product.id == lot.product_id).first()
        return {
            "type": "lot",
            "lot_id": lot.id,
            "lot_number": lot.lot_number,
            "product_id": lot.product_id,
            "product_name": product.name if product else None,
        }

    product = db.query(Product).filter(Product.barcode == code, Product.empresa_id == eid).first()
    if product:
        return {
            "type": "product",
            "product_id": product.id,
            "product_name": product.name,
        }

    raise HTTPException(404, "Código no encontrado")

# ── Excel Report ──────────────────────────────────────────────────────────────
@app.get("/reports/inventory_lots.xlsx", tags=["reports"])
def report_inventory_lots(
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
    eid: int = Depends(get_empresa_id)
):
    try:
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()

        # ── Hoja 1: Inventario por lote ──
        ws1 = wb.active
        ws1.title = "Inventario por Lote"
        ws1.append(["Producto", "Categoría", "Unidad", "Lote", "Vencimiento", "Cantidad actual", "Stock mínimo", "Código"])
        lots = (
            db.query(Lot, Product)
            .join(Product, Product.id == Lot.product_id)
            .filter(Lot.empresa_id == eid)
            .order_by(Lot.expiry_date.asc())
            .all()
        )
        for lot, product in lots:
            exp_str = lot.expiry_date.strftime("%Y-%m-%d") if lot.expiry_date else "Sin fecha"
            ws1.append([
                product.name or "", product.category or "", product.unit or "",
                lot.lot_number or "", exp_str, lot.qty_current or 0,
                product.min_stock or 0, lot.barcode or "",
            ])

        # ── Hoja 2: Movimientos (trazabilidad) ──
        ws2 = wb.create_sheet("Movimientos")
        ws2.append(["ID", "Tipo", "Producto", "Lote", "Cantidad", "Usuario", "Fecha", "Razón"])
        movements = (
            db.query(Movement)
            .options(joinedload(Movement.product), joinedload(Movement.lot))
            .filter(Movement.empresa_id == eid)
            .order_by(Movement.created_at.desc())
            .all()
        )
        for m in movements:
            ws2.append([
                m.id,
                m.type.value,
                m.product.name if m.product else "",
                m.lot.lot_number if m.lot else "",
                m.qty,
                m.user_email,
                m.created_at.strftime("%Y-%m-%d %H:%M") if m.created_at else "",
                m.reason or "",
            ])

        # ── Estilos headers ──
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for ws in [ws1, ws2]:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        logger.info(f"📊 Excel generado por {u.email}")
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=inventario_clinica.xlsx"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando reporte: {str(e)}")

# ── Email Alerts ──────────────────────────────────────────────────────────────
@app.post("/reports/email-alerts", tags=["reports"])
def send_email_alerts(
    db: Session = Depends(get_db),
    u: User = Depends(get_current_user),
    eid: int = Depends(get_empresa_id)
):
    if u.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Solo administradores")
        
    today = datetime.now().date()
    thirty_days = today + timedelta(days=30)
    
    # 1. Lotes próximos a caducar (< 30 días)
    expiring_lots = db.query(Lot).options(joinedload(Lot.product)).filter(
        Lot.empresa_id == eid,
        Lot.qty_current > 0,
        Lot.expiry_date <= thirty_days
    ).order_by(Lot.expiry_date.asc()).all()
    
    # 2. Productos bajo el stock mínimo (usando SQL puro para rapidez)
    low_stock_query = text("""
        SELECT p.name, p.min_stock, COALESCE(SUM(l.qty_current), 0) as total
        FROM products p
        LEFT JOIN lots l ON p.id = l.product_id
        WHERE p.empresa_id = :eid
        GROUP BY p.id, p.name, p.min_stock
        HAVING COALESCE(SUM(l.qty_current), 0) < p.min_stock
    """)
    res_low = db.execute(low_stock_query, {"eid": eid}).fetchall()
    
    # 3. Construir plantilla HTML
    msg_html = f"<div style='font-family: sans-serif; max-width: 600px; margin: auto; padding: 20px; border: 1px solid #e2e8f0; border-radius: 10px;'>"
    msg_html += f"<h2 style='color: #1e293b; border-bottom: 2px solid #e2e8f0; padding-bottom: 10px;'>📊 Reporte Automático de Inventario ({today})</h2>"
    
    msg_html += "<h3 style='color: #ef4444;'>🔴 Riesgo de Desabastecimiento (Bajo Mínimo)</h3><ul style='list-style: none; padding: 0;'>"
    if not res_low:
        msg_html += "<li style='padding: 10px; background: #f0fdf4; color: #15803d; border-radius: 5px;'>✓ Todos los productos tienen stock suficiente.</li>"
    for r in res_low:
        msg_html += f"<li style='padding: 10px; background: #fef2f2; margin-bottom: 5px; border-radius: 5px; border-left: 4px solid #ef4444;'>"
        msg_html += f"<b>{r[0]}</b> <br> Quedan: <b>{r[2]}</b> (El mínimo permitido es {r[1]})</li>"
    msg_html += "</ul>"
    
    msg_html += "<h3 style='color: #f59e0b; margin-top: 30px;'>⚠️ Lotes próximos a Caducar (<= 30 días)</h3><ul style='list-style: none; padding: 0;'>"
    if not expiring_lots:
        msg_html += "<li style='padding: 10px; background: #f0fdf4; color: #15803d; border-radius: 5px;'>✓ Ningún lote caduca pronto.</li>"
    for l in expiring_lots:
        product_name = l.product.name if l.product else 'Desconocido'
        msg_html += f"<li style='padding: 10px; background: #fffbeb; margin-bottom: 5px; border-radius: 5px; border-left: 4px solid #f59e0b;'>"
        msg_html += f"<b>{product_name}</b> (Lote: {l.lot_number}) <br> Vence: <b>{l.expiry_date}</b> | Stock Actual: {l.qty_current}</li>"
    msg_html += "</ul></div>"
    
    # 4. Intentar enviar correo
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    
    smtp_server = os.getenv("SMTP_SERVER")
    smtp_user = os.getenv("SMTP_USERNAME")
    smtp_pass = os.getenv("SMTP_PASSWORD")
    to_email = os.getenv("ADMIN_EMAIL", u.email)
    
    if smtp_server and smtp_user and smtp_pass:
        try:
            msg = MIMEMultipart("alternative")
            msg["Subject"] = f"Alerta de Clínica: Inventario - {today}"
            msg["From"] = smtp_user
            msg["To"] = to_email
            msg.attach(MIMEText(msg_html, "html"))
            
            with smtplib.SMTP_SSL(smtp_server, 465) as server:
                server.login(smtp_user, smtp_pass)
                server.sendmail(smtp_user, to_email, msg.as_string())
                
            logger.info(f"📧 Correo de alerta enviado a {to_email}")
            return {"message": f"Reporte enviado con éxito a {to_email}"}
        except Exception as e:
            logger.error(f"Error enviando alertas SMTP: {e}")
            raise HTTPException(500, f"Error SMTP de envío: {str(e)}")
    else:
        logger.warning(f"SIMULACRO CORREO (SMTP no configurado en .env):\n{msg_html}")
        return {"message": "Analizado con éxito. Te faltó configurar tu correo SMTP en Render (.env) para recibir el email real. Revisa la consola para ver el simulacro."}

# ── Health ────────────────────────────────────────────────────────────────────
@app.get("/health", tags=["system"])
def health_check():
    return {"status": "ok", "version": "1.2.0"}

@app.get("/", tags=["system"])
def root():
    return {"message": "API Inventario Clínica", "version": "1.2.0", "docs": "/docs"}
